from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(r"C:\Users\yshvadro\OneDrive - Jaggaer\Customer Success - General\Value Articulation")
OUT = ROOT / "_extracted_value_articulation.json"

NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
}

KEYWORDS = [
    "savings",
    "saving",
    "cost",
    "hard",
    "soft",
    "avoidance",
    "reduction",
    "reduced",
    "discount",
    "rebate",
    "compliance",
    "cycle time",
    "automation",
    "efficiency",
    "fte",
    "headcount",
    "value",
    "roi",
    "spend",
    "po",
    "invoice",
    "sourcing",
    "contract",
]


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def hit_lines(text: str) -> list[str]:
    lines = []
    for raw in text.splitlines():
        line = normalize(raw)
        if not line:
            continue
        low = line.lower()
        if any(k in low for k in KEYWORDS):
            lines.append(line)
    return lines[:400]


def extract_docx(path: Path) -> dict:
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    paras = []
    for para in root.findall(".//w:p", NS):
        texts = [t.text or "" for t in para.findall(".//w:t", NS)]
        joined = normalize("".join(texts))
        if joined:
            paras.append(joined)
    full_text = "\n".join(paras)
    return {"type": "docx", "hits": hit_lines(full_text), "sample": paras[:250]}


def extract_pptx(path: Path) -> dict:
    slides = []
    all_lines = []
    with zipfile.ZipFile(path) as zf:
        slide_names = sorted(
            name
            for name in zf.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        for name in slide_names:
            root = ET.fromstring(zf.read(name))
            texts = [normalize(t.text or "") for t in root.findall(".//a:t", NS)]
            texts = [t for t in texts if t]
            slide_text = "\n".join(texts)
            slide_hits = hit_lines(slide_text)
            slides.append(
                {
                    "slide": name.split("/")[-1].replace(".xml", ""),
                    "hits": slide_hits[:40],
                    "sample": texts[:80],
                }
            )
            all_lines.extend(texts)
    return {"type": "pptx", "slides": slides, "hits": hit_lines("\n".join(all_lines))}


def extract_xlsx(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    result_sheets = []
    all_lines = []
    formula_rows = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows():
            cells = []
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                value_str = str(value)
                if len(value_str) > 500:
                    value_str = value_str[:500]
                cells.append(f"{cell.coordinate}: {normalize(value_str)}")
                if isinstance(value, str) and value.startswith("="):
                    low_formula = value.lower()
                    if any(k in low_formula for k in ["sum", "if", "count", "xlookup", "vlookup", "average", "round"]):
                        formula_rows.append(f"{ws.title}!{cell.coordinate}: {value}")
            if cells:
                rows.append(" | ".join(cells))
        joined = "\n".join(rows)
        hits = hit_lines(joined)
        result_sheets.append(
            {
                "sheet": ws.title,
                "hits": hits[:80],
                "sample_rows": rows[:120],
            }
        )
        all_lines.extend(rows)
    return {
        "type": "xlsx",
        "sheets": result_sheets,
        "hits": hit_lines("\n".join(all_lines)),
        "formula_samples": formula_rows[:400],
    }


def extract_pdf(path: Path) -> dict:
    reader = PdfReader(str(path))
    pages = []
    all_text = []
    for idx, page in enumerate(reader.pages, start=1):
        text = normalize(page.extract_text() or "")
        if not text:
            continue
        pages.append({"page": idx, "hits": hit_lines(text), "sample": text[:4000]})
        all_text.append(text)
    return {"type": "pdf", "pages": pages, "hits": hit_lines("\n".join(all_text))}


def main() -> None:
    extracted = {}
    files = sorted(
        p for p in ROOT.rglob("*")
        if p.is_file() and p.suffix.lower() in {".docx", ".pptx", ".xlsx", ".pdf"}
    )
    for path in files:
        rel = str(path.relative_to(ROOT))
        try:
            if path.suffix.lower() == ".docx":
                extracted[rel] = extract_docx(path)
            elif path.suffix.lower() == ".pptx":
                extracted[rel] = extract_pptx(path)
            elif path.suffix.lower() == ".xlsx":
                extracted[rel] = extract_xlsx(path)
            elif path.suffix.lower() == ".pdf":
                extracted[rel] = extract_pdf(path)
        except Exception as exc:  # pragma: no cover
            extracted[rel] = {"error": str(exc)}
    OUT.write_text(json.dumps(extracted, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
